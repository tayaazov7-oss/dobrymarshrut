import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(__file__)
FRONT_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
DB_PATH = os.path.join(BASE_DIR, "db.sqlite")
XLSX_PATH = os.path.join(BASE_DIR, "database.xlsx")


# схема таблицы NKO — должна совпадать с server.py
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS nko (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            description TEXT,
            volunteers TEXT,
            phone TEXT,
            address TEXT,
            logo TEXT,
            website TEXT,
            albums TEXT,
            filters TEXT,
            city TEXT,
            lat REAL,
            lng REAL,
            status TEXT NOT NULL DEFAULT 'approved',
            created_by_user_id INTEGER,
            created_at TEXT NOT NULL
        );
    """)

    conn.commit()
    conn.close()


def detect_columns(header_row):
    """
    Определяем, в какой колонке какое поле.
    Ориентируемся по русским названиям из Excel.
    """

    mapping = {
        "name": None,
        "category": None,
        "description": None,
        "volunteers": None,
        "phone": None,
        "address": None,
        "logo": None,
        "website": None,
        "albums": None,
        "filters": None,
        "city": None,
        "lat": None,
        "lng": None,
    }

    for idx, cell in enumerate(header_row, start=1):
        if cell.value is None:
            continue
        title = str(cell.value).strip().lower()

        # Название
        if any(key in title for key in ["название", "наименование", "нко"]):
            mapping["name"] = idx

        # Категория / направление
        elif "категор" in title or "направлен" in title:
            mapping["category"] = idx

        # Описание
        elif "описан" in title:
            mapping["description"] = idx

        # Функционал волонтёров
        elif "волонтер" in title or "волонтёр" in title:
            mapping["volunteers"] = idx

        # Телефон
        elif "телефон" in title or "тел." in title:
            mapping["phone"] = idx

        # Адрес
        elif "адрес" in title:
            mapping["address"] = idx

        # Лого
        elif "лого" in title or "логотип" in title:
            mapping["logo"] = idx

        # Сайт / соцсети
        elif "сайт" in title or "соцсет" in title or "соцсет" in title:
            mapping["website"] = idx

        # Альбомы / мероприятия
        elif "альбом" in title or "меропр" in title:
            mapping["albums"] = idx

        # Фильтры
        elif "фильтр" in title or "простые фильтры" in title:
            mapping["filters"] = idx

        # Город
        elif "город" in title:
            mapping["city"] = idx

        # Координаты (если есть)
        elif "широт" in title or "latitude" in title or "lat" == title:
            mapping["lat"] = idx
        elif "долгот" in title or "longitude" in title or "lng" == title:
            mapping["lng"] = idx

    return mapping


def parse_excel_to_db(xlsx_path=XLSX_PATH, db_path=DB_PATH):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel-файл не найден: {xlsx_path}")

    init_db()

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # берем первый лист

    rows = list(ws.iter_rows())
    if not rows:
        print("Пустой Excel-файл.")
        return

    header_row = rows[0]
    mapping = detect_columns(header_row)

    if mapping["name"] is None:
        raise RuntimeError("Не удалось найти колонку с названием НКО (\"Название\")")

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    inserted = 0

    for row in rows[1:]:
        # проверяем, есть ли название
        name_cell = row[mapping["name"] - 1]
        name = (str(name_cell.value).strip()
                if name_cell.value is not None else "")
        if not name:
            continue  # пропускаем пустые строки

        def get_value(key):
            col = mapping.get(key)
            if not col:
                return ""
            cell = row[col - 1]
            return "" if cell.value is None else str(cell.value).strip()

        category = get_value("category")
        description = get_value("description")
        volunteers = get_value("volunteers") or description
        phone = get_value("phone")
        address = get_value("address")
        logo = get_value("logo")
        website = get_value("website")
        albums = get_value("albums")
        filters = get_value("filters")
        city = get_value("city")

        lat = None
        lng = None
        # если в таблице уже есть координаты, пробуем считать
        if mapping.get("lat"):
            try:
                cell = row[mapping["lat"] - 1]
                lat = float(cell.value) if cell.value is not None else None
            except (ValueError, TypeError):
                lat = None

        if mapping.get("lng"):
            try:
                cell = row[mapping["lng"] - 1]
                lng = float(cell.value) if cell.value is not None else None
            except (ValueError, TypeError):
                lng = None

        cur.execute("""
            INSERT INTO nko (
                name, category, description, volunteers,
                phone, address, logo, website, albums, filters,
                city, lat, lng, status, created_by_user_id, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            name, category, description, volunteers,
            phone, address, logo, website, albums, filters,
            city, lat, lng, "approved", None, datetime.utcnow().isoformat()
        ))

        inserted += 1

    conn.commit()
    conn.close()

    print(f"Импортировано записей НКО: {inserted}")


if __name__ == "__main__":
    parse_excel_to_db()